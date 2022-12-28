import os
import pathlib
from datetime import date
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter.ttk import Combobox

import background as background
import openpyxl
import xlrd
from PIL import Image, ImageTk
from openpyxl.workbook import Workbook

background = '#06283D'
framebg = '#EDEDED'
framefg = '#06283D'

root = Tk()  # create object
root.geometry("1250x700+210+100")  # geometry("width x height")
root.title('Student Registration ')  # title of the page
root.config(bg=background)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = 'Registration No.'
    sheet['B1'] = 'Name'
    sheet['C1'] = 'Class'
    sheet['D1'] = 'Gender'
    sheet['E1'] = 'DOB'
    sheet['F1'] = 'Date of Registration'
    sheet['G1'] = 'Religion'
    sheet['H1'] = 'Skill'
    sheet['I1'] = 'Father Name'
    sheet['J1'] = 'Mother Name'
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"

    file.save('Student_data.xlsx')


# gender

def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = 'Male'


    else:
        gender = 'Female'


# showimage

def showimage():
    global fileimages
    global img
    fileimages = filedialog.askopenfilename(initialdir=os.getcwd(), title='select image file',
                                            filetypes=(("jpg File", "*.jpg"), ('PNG File', '*.png'),
                                                       ('ALL files', '*.txt')))

    img = (Image.open(fileimages))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


### Search

def search():
    global reg_number
    text = Search.get()  # taking inpute from entry box

    Clear()  # to clear all  data

    saveButton.config(state='disable')  # after clicking on sesrch

    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]

            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    try:
        print(str(name))

    except:
        messagebox.showerror('Invalid', 'Invalid registration number!!!')

    # reg_no_position showing link A2,A3,......An
    # but reg_number just showing number after A2 like 2,3,.....,n

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4 == 'Female':
        R2.select()

    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_name.set(x9)
    M_name.get(x10)
    F_occupation.get(x11)
    M_occupation.get(x12)

    img = (Image.open('Student Image/' + str(x1) + '.jpg'))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


###Update

def Update():
    global reg_number
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender()
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_name.get()
    mothername = M_name.get()
    F1 = F_occupation.get()
    M1 = M_occupation.get()

    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=N1)
    sheet.cell(column=3, row=int(reg_number), value=C1)
    sheet.cell(column=4, row=int(reg_number), value=G1)
    sheet.cell(column=5, row=int(reg_number), value=D2)
    sheet.cell(column=6, row=int(reg_number), value=D1)
    sheet.cell(column=7, row=int(reg_number), value=Re1)
    sheet.cell(column=8, row=int(reg_number), value=S1)
    sheet.cell(column=9, row=int(reg_number), value=fathername)
    sheet.cell(column=10, row=int(reg_number), value=mothername)
    sheet.cell(column=11, row=int(reg_number), value=F1)
    sheet.cell(column=12, row=int(reg_number), value=M1)

    file.save(r'Student_date.xlsx')

    try:
        img.save('Student Images/' + str(R1) + '.jpg')

    except:
        pass
    messagebox.showinfo('Update', 'update Sucessfully!!')

    Clear()


########################################### registration no #######################
# now each time we have to enter registration no .
# lets design automatic registration no. entry system
# it is  created to automatic enter registration no .

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)

    except:
        Registration.set("1")


# Clear

def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_name.set('')
    M_name.set('')
    F_occupation.set('')
    M_occupation.set('')
    Class.set('Select Class')

    registration_no()

    saveButton.config(state='normal')

    img1 = PhotoImage(file='Image/img4.png')
    lbl.config(image=img1)
    lbl.image = img1

    img = ''


###########################################Save

def Save():
    global G1
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender

    except:
        messagebox.showerror('error', 'Select Gender!')

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_name.get()
    mothername = M_name.get()
    F1 = F_occupation.get()
    M1 = M_occupation.get()

    if N1 == '' or C1 == 'Select Class' or D2 == '' or Re1 == '' or S1 == '' or fathername == '' or mothername == '' or F1 == '' or M1 == '':
        messagebox.showerror('error', 'Few Date is missing!')
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=fathername)
        sheet.cell(column=10, row=sheet.max_row, value=mothername)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)

        file.save(r'Student_date.xlsx')
        try:
            img.save('Student Image/' + str(R1) + '.jpg')

        except:
            messagebox.showerror("info", 'Picture is not available!!!!')

        messagebox.showinfo('info', 'SucessFully data entered!!!!')

        Clear()  # clear entry box and image section

        registration_no()  # it will recheck registration no. and reissue new


# top frames

Label(root, text='Satyam ', width=10, height=3, bg='#f0687c', anchor='e').pack(side=TOP, fill=X)
Label(root, text='STUDENT REGISTRATION', width=10, height=2, bg='#c36464', fg='#fff', font='arial 20 bold').pack(
    side=TOP, fill=X)

# search box to update

Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font='arial 20').place(x=820, y=70)
imageicon3 = PhotoImage(file='Image/img1.png')
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, bg='#68ddfa', font='arial 13 bold', command=search)
Srch.place(x=1070, y=69)

imageicon4 = PhotoImage(file='Image/img2.png')
Update_button = Button(root, image=imageicon4, bg='#c36464', command=Update)
Update_button.place(x=110, y=62)

# Registration and Date

Label(root, text='Registration No : ', font='arial 13', fg=framebg, bg=background).place(x=30, y=150)
Label(root, text='Date: ', font='arial 13', fg=framebg, bg=background).place(x=500, y=150)

# Label(root,text=' : ',font='arial 13',fg=framebg,bg=background).place(x=30,y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font='arial 10')
reg_entry.place(x=160, y=155)

# registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = Entry(root, textvariable=Date, width=15, font='arial 10')
date_entry.place(x=550, y=155)

Date.set(d1)

# Student details

obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text='Full Name : ', font='arial 13', bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text='Date of Birth : ', font='arial 13', bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text='Gender : ', font='arial 13', bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text='Class : ', font='arial 13', bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text='Religion : ', font='arial 13', bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text='Skills : ', font='arial 13', bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font='arial 10')
name_entry.place(x=160, y=55)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font='arial 10')
dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text='Male', variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text='Female', variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font='arial 10')
religion_entry.place(x=630, y=100)

Skill = StringVar()
skill_entry = Entry(obj, textvariable=Skill, width=20, font='arial 10')
skill_entry.place(x=630, y=150)

Class = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9'], font='Roboto 10', width=17, state='r')
Class.place(x=630, y=50)
Class.set('Select Class')

# Parents details

obj = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, height=250, relief=GROOVE)
obj.place(x=30, y=470)

Label(obj, text="Father's name ", font='arial 13', bg=framebg, fg=framefg).place(x=30, y=50)

F_name = StringVar()
f_entry = Entry(obj, textvariable=F_name, width=20, font='arial 10')
f_entry.place(x=160, y=50)

Label(obj, text="Occupation", font='arial 13', bg=framebg, fg=framefg).place(x=30, y=100)

F_occupation = StringVar()
f_entry = Entry(obj, textvariable=F_occupation, width=20, font='arial 10')
f_entry.place(x=160, y=100)

Label(obj, text="Mother's name ", font='arial 13', bg=framebg, fg=framefg).place(x=500, y=50)

M_name = StringVar()
m_entry = Entry(obj, textvariable=M_name, width=20, font='arial 10')
m_entry.place(x=630, y=50)

Label(obj, text="Occupation", font='arial 13', bg=framebg, fg=framefg).place(x=500, y=100)

M_occupation = StringVar()
m_entry = Entry(obj, textvariable=M_occupation, width=20, font='arial 10')
m_entry.place(x=630, y=100)

# image

f = Frame(root, bd=3, bg='black', width=180, height=180, relief=GROOVE)
f.place(x=1000, y=192)

img = PhotoImage(file='Image/img4.png')
lbl = Label(f, bg='black', image=img)
lbl.place(x=0, y=0)

# button


Button(root, text='Upload', width=17, height=2, font='arial 12 bold', bg='lightblue', command=showimage).place(x=1000,
                                                                                                               y=380)

saveButton = Button(root, text='Save', width=17, height=2, font='arial 12 bold', bg='lightblue', command=Save)
saveButton.place(x=1000, y=450)

resetButton = Button(root, text='Reset', width=17, height=2, font='arial 12 bold', bg='lightblue', command=Clear)
resetButton.place(x=1000, y=520)


# exit
def Exit():
    root.destroy()


exitButton = Button(root, text='Exit', width=17, height=2, font='arial 12 bold', bg='lightblue', command=Exit)

exitButton.place(x=1000, y=590)









root.mainloop()
